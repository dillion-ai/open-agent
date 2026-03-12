"""Recalculate all formulas in an Excel file using LibreOffice and report errors."""

import json
import os
import subprocess
import sys

from office.soffice import find_soffice, get_user_profile_dir

EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def recalc(file_path: str, timeout: int = 60) -> dict:
    file_path = os.path.abspath(file_path)
    if not os.path.isfile(file_path):
        return {"status": "error", "message": f"File not found: {file_path}"}

    soffice = find_soffice()
    profile = get_user_profile_dir()

    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "gen"

    # Use LibreOffice to open, recalculate, and save
    outdir = os.path.dirname(file_path)
    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "--nolockcheck",
        "--calc",
        f"-env:UserInstallation=file://{profile}",
        "--convert-to", "xlsx",
        "--outdir", outdir,
        file_path,
    ]

    try:
        subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, env=env)
    except subprocess.TimeoutExpired:
        return {"status": "error", "message": f"LibreOffice timed out after {timeout}s"}

    # Now scan for errors using openpyxl
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}

    wb = load_workbook(file_path, data_only=True)
    errors = {}
    total_formulas = 0
    total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val in EXCEL_ERRORS:
                    total_errors += 1
                    if val not in errors:
                        errors[val] = {"count": 0, "locations": []}
                    errors[val]["count"] += 1
                    errors[val]["locations"].append(f"{sheet_name}!{cell.coordinate}")

    # Count formulas from the non-data-only workbook
    wb2 = load_workbook(file_path)
    for sheet_name in wb2.sheetnames:
        ws = wb2[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    total_formulas += 1

    result = {
        "status": "errors_found" if total_errors > 0 else "success",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
    }
    if errors:
        result["error_summary"] = errors

    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <excel_file> [timeout_seconds]", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    result = recalc(path, timeout)
    print(json.dumps(result, indent=2))
